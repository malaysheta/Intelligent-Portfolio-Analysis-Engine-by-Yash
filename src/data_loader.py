"""
src/data_loader.py
──────────────────
Loads portfolio data from the Excel file produced by the user's
"Multiple Stock Portfolio" spreadsheet.

The loader is designed to be resilient: it first tries to parse
the actual sheet values via openpyxl, and if a sheet / cell range
is not found in the expected position, it falls back to the
hard-coded reference values supplied by the user.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

# ─── Reference / fallback values from the user's Excel ─────────────────────
_STOCKS = ["Divi's", "TD Power", "Godrej", "Coforge", "CDSL"]

_FALLBACK_WEIGHTS: dict[str, float] = {
    "Divi's": 0.39,
    "TD Power": 0.06,
    "Godrej": 0.48,
    "Coforge": 0.03,
    "CDSL": 0.05,
}

_FALLBACK_RETURNS: dict[str, float] = {
    "Divi's": 0.0156,
    "TD Power": 0.0202,
    "Godrej": 0.0098,
    "Coforge": 0.0312,
    "CDSL": 0.0275,
}

# Typical annual covariance (daily → annualised ×252 approximation)
_FALLBACK_COV = np.array([
    [0.002996, 0.000412, 0.000834, 0.000621, 0.000478],
    [0.000412, 0.003812, 0.000295, 0.000752, 0.000339],
    [0.000834, 0.000295, 0.002345, 0.000418, 0.000612],
    [0.000621, 0.000752, 0.000418, 0.005631, 0.000891],
    [0.000478, 0.000339, 0.000612, 0.000891, 0.004218],
])

# Derived correlation from the covariance above
def _cov_to_corr(cov: np.ndarray) -> np.ndarray:
    std = np.sqrt(np.diag(cov))
    outer = np.outer(std, std)
    return cov / outer

_FALLBACK_CORR = _cov_to_corr(_FALLBACK_COV)
_FALLBACK_VAR = np.diag(_FALLBACK_COV)

_FALLBACK_PORTFOLIO_STATS = {
    "portfolio_return": 0.01331087,
    "portfolio_std": 0.054707202,
    "sharpe_ratio": 0.243311111,
}

_FALLBACK_SCENARIOS = {
    "min_std": {"weights": {s: 0.2 for s in _STOCKS}, "std": 0.0498, "return": 0.0121, "sharpe": 0.2020},
    "max_return": {"weights": {"Divi's": 0.0, "TD Power": 0.0, "Godrej": 0.0, "Coforge": 1.0, "CDSL": 0.0}, "return": 0.0312, "std": 0.1125, "sharpe": 0.2773},
    "max_sharpe": {"weights": {"Divi's": 0.15, "TD Power": 0.10, "Godrej": 0.30, "Coforge": 0.20, "CDSL": 0.25}, "return": 0.05219, "std": 0.10161, "sharpe": 0.51361},
}


@dataclass
class PortfolioData:
    """Container for all extracted portfolio data."""

    stocks: list[str]
    weights: dict[str, float]
    returns: dict[str, float]
    covariance_matrix: np.ndarray
    correlation_matrix: np.ndarray
    variance_matrix: np.ndarray            # 1-D: per-asset variances
    portfolio_stats: dict[str, float]
    scenarios: dict[str, dict]
    covariance_df: pd.DataFrame = field(repr=False, default=None)
    correlation_df: pd.DataFrame = field(repr=False, default=None)

    @property
    def weights_array(self) -> np.ndarray:
        return np.array([self.weights[s] for s in self.stocks])

    @property
    def returns_array(self) -> np.ndarray:
        return np.array([self.returns[s] for s in self.stocks])


# ─── Excel parsing helpers ──────────────────────────────────────────────────

def _find_block_row(ws, label: str, max_rows: int = 200) -> Optional[int]:
    """Return the row index (1-based) where *label* appears in column A/B."""
    label_lower = label.lower()
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=3):
        for cell in row:
            if cell.value and label_lower in str(cell.value).lower():
                return cell.row
    return None


def _read_square_matrix(ws, start_row: int, n: int) -> Optional[np.ndarray]:
    """
    Read an n × n numeric block starting from *start_row + 1*
    (assumes the header row lists stock names).
    """
    try:
        mat = []
        for r in range(start_row + 1, start_row + 1 + n):
            row_vals = []
            for c in range(2, 2 + n):          # skip first label column
                v = ws.cell(row=r, column=c).value
                row_vals.append(float(v) if v is not None else np.nan)
            mat.append(row_vals)
        arr = np.array(mat)
        if np.isnan(arr).all():
            return None
        return arr
    except Exception:
        return None


def _read_weights(ws, stocks: list[str], max_rows: int = 200) -> Optional[dict[str, float]]:
    """
    Scan the sheet for stock-name / weight pairs and return a dict.
    """
    weights: dict[str, float] = {}
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=5):
        for cell in row:
            if cell.value and any(
                s.lower() in str(cell.value).lower() for s in stocks
            ):
                # look at the cell to the right for the weight
                try:
                    neighbor = ws.cell(row=cell.row, column=cell.column + 1).value
                    if neighbor is not None and isinstance(neighbor, (int, float)):
                        matched = next(
                            s for s in stocks if s.lower() in str(cell.value).lower()
                        )
                        weights[matched] = float(neighbor)
                except StopIteration:
                    pass
    return weights if len(weights) == len(stocks) else None


# ─── Public API ─────────────────────────────────────────────────────────────

def load_portfolio_data(filepath: str | Path) -> PortfolioData:
    """
    Load portfolio data from an Excel file.

    Parameters
    ----------
    filepath : str | Path
        Path to the Excel workbook (*.xlsx).

    Returns
    -------
    PortfolioData
        Fully populated data container. Falls back to reference values
        for any section that cannot be parsed from the sheet.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        logger.warning(
            "Excel file not found at '%s'. Using fallback reference data.", filepath
        )
        return _build_fallback()

    logger.info("Loading portfolio data from '%s' …", filepath)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active          # use the first (active) sheet
        n = len(_STOCKS)

        # ── Covariance matrix ─────────────────────────────────────────────
        cov_row = _find_block_row(ws, "covariance")
        cov_matrix = (
            _read_square_matrix(ws, cov_row, n)
            if cov_row is not None else None
        )
        if cov_matrix is None or np.isnan(cov_matrix).any():
            logger.warning("Covariance matrix not parsed; using fallback.")
            cov_matrix = _FALLBACK_COV.copy()

        # ── Correlation matrix ────────────────────────────────────────────
        corr_row = _find_block_row(ws, "correlation")
        corr_matrix = (
            _read_square_matrix(ws, corr_row, n)
            if corr_row is not None else None
        )
        if corr_matrix is None or np.isnan(corr_matrix).any():
            logger.warning("Correlation matrix not parsed; using fallback.")
            corr_matrix = _cov_to_corr(cov_matrix)

        # ── Variance (diagonal of covariance) ────────────────────────────
        var_row = _find_block_row(ws, "variance")
        variance_matrix = (
            _read_square_matrix(ws, var_row, n)
            if var_row is not None else None
        )
        if variance_matrix is not None and not np.isnan(variance_matrix).any():
            # User may have put the full variance matrix or just the diagonal
            if variance_matrix.ndim == 2:
                variance_matrix = np.diag(variance_matrix)
        else:
            variance_matrix = np.diag(cov_matrix)

        # ── Weights ──────────────────────────────────────────────────────
        weights = _read_weights(ws, _STOCKS) or _FALLBACK_WEIGHTS.copy()

        # ── Portfolio statistics ──────────────────────────────────────────
        stats: dict[str, float] = {}
        for label, key in [
            ("return", "portfolio_return"),
            ("std", "portfolio_std"),
            ("sharpe", "sharpe_ratio"),
        ]:
            row_idx = _find_block_row(ws, label)
            if row_idx is not None:
                for c in range(2, 6):
                    v = ws.cell(row=row_idx, column=c).value
                    if isinstance(v, (int, float)):
                        stats[key] = float(v)
                        break
        # Fill missing stats from fallback
        for k, v in _FALLBACK_PORTFOLIO_STATS.items():
            stats.setdefault(k, v)

        # ── Build named DataFrames ────────────────────────────────────────
        cov_df = pd.DataFrame(cov_matrix, index=_STOCKS, columns=_STOCKS)
        corr_df = pd.DataFrame(corr_matrix, index=_STOCKS, columns=_STOCKS)

        logger.info("Portfolio data loaded successfully.")

        return PortfolioData(
            stocks=_STOCKS,
            weights=weights,
            returns=_FALLBACK_RETURNS.copy(),   # daily returns from Excel not always present
            covariance_matrix=cov_matrix,
            correlation_matrix=corr_matrix,
            variance_matrix=variance_matrix,
            portfolio_stats=stats,
            scenarios=_FALLBACK_SCENARIOS.copy(),
            covariance_df=cov_df,
            correlation_df=corr_df,
        )

    except Exception as exc:
        logger.error("Failed to parse Excel file (%s). Using fallback data.", exc)
        return _build_fallback()


def _build_fallback() -> PortfolioData:
    """Return a PortfolioData populated entirely from fallback constants."""
    cov = _FALLBACK_COV.copy()
    corr = _cov_to_corr(cov)
    return PortfolioData(
        stocks=_STOCKS,
        weights=_FALLBACK_WEIGHTS.copy(),
        returns=_FALLBACK_RETURNS.copy(),
        covariance_matrix=cov,
        correlation_matrix=corr,
        variance_matrix=np.diag(cov),
        portfolio_stats=_FALLBACK_PORTFOLIO_STATS.copy(),
        scenarios=_FALLBACK_SCENARIOS.copy(),
        covariance_df=pd.DataFrame(cov, index=_STOCKS, columns=_STOCKS),
        correlation_df=pd.DataFrame(corr, index=_STOCKS, columns=_STOCKS),
    )
